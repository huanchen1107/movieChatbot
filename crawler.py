import requests
import urllib3
from bs4 import BeautifulSoup
import re
import time
import sys
import io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from db import init_db, save_movies, load_movies

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

BASE_URL = "https://ssr1.scrape.center"
OUTPUT_DIR = Path(__file__).parent / "output"
POSTER_DIR = OUTPUT_DIR / "posters"
OUTPUT_DIR.mkdir(exist_ok=True)
POSTER_DIR.mkdir(exist_ok=True)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

COLUMNS = ["id", "name", "categories", "country", "duration", "release_date", "score"]
POSTER_CELL_WIDTH = 18
POSTER_CELL_HEIGHT = 260
ROW_HEIGHT = 260
THUMB_HEIGHT = 240


def fetch(url, retries=3):
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=15, verify=False)
            resp.raise_for_status()
            resp.encoding = "utf-8"
            return resp
        except requests.RequestException as e:
            print(f"  [!] Attempt {attempt+1}/{retries} failed: {e}")
            if attempt < retries - 1:
                time.sleep(2)
    return None


def download_image(url, filepath):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30, verify=False)
        resp.raise_for_status()
        with open(filepath, "wb") as f:
            f.write(resp.content)
        return True
    except Exception as e:
        print(f"  [!] Download poster failed: {e}")
        return False


def parse_movie_card(card):
    name_el = card.select_one(".name h2")
    name = name_el.get_text(strip=True) if name_el else ""

    detail_link = card.select_one("a[href*='/detail/']")
    movie_id = ""
    if detail_link:
        m = re.search(r"/detail/(\d+)", detail_link["href"])
        if m:
            movie_id = m.group(1)

    categories = [btn.get_text(strip=True) for btn in card.select(".category span")]

    cover_el = card.select_one(".cover")
    cover = cover_el.get("src", "") if cover_el else ""

    score_el = card.select_one(".score")
    score = score_el.get_text(strip=True) if score_el else ""

    info_spans = card.select(".info span")
    info_parts = [s.get_text(strip=True) for s in info_spans if s.get_text(strip=True) != "/"]
    country = info_parts[0] if len(info_parts) >= 1 else ""
    duration = info_parts[1] if len(info_parts) >= 2 else ""
    release_date = info_parts[2] if len(info_parts) >= 3 else ""

    return {
        "id": movie_id,
        "name": name,
        "categories": categories,
        "country": country,
        "duration": duration,
        "release_date": release_date,
        "score": score,
        "cover": cover,
    }


def crawl_page(page_num):
    url = f"{BASE_URL}/page/{page_num}" if page_num > 1 else BASE_URL
    print(f"[*] Crawling: {url}")
    resp = fetch(url)
    if not resp:
        return [], 0

    soup = BeautifulSoup(resp.text, "html.parser")
    cards = soup.select("#index .el-card.item")
    movies = [parse_movie_card(card) for card in cards]

    total_text = soup.select_one(".el-pagination__total")
    total = 0
    if total_text:
        m = re.search(r"(\d+)", total_text.get_text())
        if m:
            total = int(m.group(1))

    return movies, total


def crawl_all():
    print("=" * 60)
    print("  Scrape.center Movie Crawler")
    print("=" * 60)

    movies, total = crawl_page(1)
    if not movies or total == 0:
        print("[!] Failed to fetch page 1. Exiting.")
        return []

    print(f"[+] Found {total} movies total.")

    page_num = 2
    per_page = len(movies)
    total_pages = (total + per_page - 1) // per_page

    while page_num <= total_pages:
        new_movies, _ = crawl_page(page_num)
        if not new_movies:
            print(f"[!] Page {page_num} fetch failed, stopping.")
            break
        movies.extend(new_movies)
        page_num += 1
        time.sleep(1)

    return movies


def download_posters(movies):
    print(f"\n[*] Downloading {len(movies)} posters...")
    for i, m in enumerate(movies, 1):
        cover_url = m.get("cover", "")
        if not cover_url:
            continue

        ext = ".jpg"
        if cover_url.rfind(".") > cover_url.rfind("/"):
            ext = cover_url[cover_url.rfind("."):]
            if len(ext) > 5:
                ext = ".jpg"

        poster_path = POSTER_DIR / f"{m['id']}{ext}"
        if poster_path.exists():
            m["poster_file"] = str(poster_path)
            continue

        print(f"  [{i}/{len(movies)}] {m['name'][:30]}...")
        if download_image(cover_url, poster_path):
            m["poster_file"] = str(poster_path)
        time.sleep(0.5)

    return movies


def save_to_db(movies):
    save_movies(movies)
    db_path = OUTPUT_DIR / "movie.db"
    print(f"\n[+] Saved {len(movies)} movies to {db_path}")


def save_csv(movies):
    import csv
    csv_path = OUTPUT_DIR / "movies.csv"
    fields = COLUMNS + ["cover", "poster_file"]
    rows = []
    for m in movies:
        cats_str = " / ".join(m["categories"]) if isinstance(m["categories"], list) else m.get("categories", "")
        rows.append({
            "id": m["id"],
            "name": m["name"],
            "categories": cats_str,
            "country": m.get("country", ""),
            "duration": m.get("duration", ""),
            "release_date": m.get("release_date", ""),
            "score": m.get("score", ""),
            "cover": m.get("cover", ""),
            "poster_file": m.get("poster_file", ""),
        })
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"[+] Saved {len(rows)} rows to {csv_path}")


def save_excel(movies):
    excel_path = OUTPUT_DIR / "movies.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "电影列表"

    excel_headers = ["ID", "海报", "电影名称", "类型", "国家/地区", "片长", "上映日期", "评分"]
    col_indexes = {h: i + 1 for i, h in enumerate(excel_headers)}

    header_fill = {"fill": None}
    header_font = None
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        body_font = Font(name="Microsoft YaHei", size=11)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
    except ImportError:
        header_fill = None
        header_font = None

    for col_idx, header in enumerate(excel_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if header_fill:
            cell.fill = header_fill
        if header_font:
            cell.font = header_font
        if header in ["ID", "评分"]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif header == "海报":
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 30

    col_widths = {
        "A": 8,
        "B": POSTER_CELL_WIDTH,
        "C": 48,
        "D": 26,
        "E": 22,
        "F": 12,
        "G": 18,
        "H": 8,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row_idx, m in enumerate(movies, 2):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT

        values = [
            (1, int(m.get("id", 0) or 0)),
            (3, m.get("name", "")),
            (4, m.get("categories", "")),
            (5, m.get("country", "")),
            (6, m.get("duration", "")),
            (7, m.get("release_date", "")),
            (8, m.get("score", "")),
        ]
        for col, val in values:
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col in [1, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        poster_file = m.get("poster_file", "")
        if poster_file:
            poster_path = Path(poster_file)
            if poster_path.exists():
                try:
                    with PILImage.open(poster_path) as img:
                        w, h = img.size
                        ratio = THUMB_HEIGHT / h
                        new_w = int(w * ratio)
                        new_h = THUMB_HEIGHT

                    img_obj = XLImage(poster_path)
                    img_obj.width = new_w
                    img_obj.height = new_h

                    cell = ws.cell(row=row_idx, column=2)
                    col_letter = "B"
                    cell_width_px = int(POSTER_CELL_WIDTH * 7)
                    col_offset = max(0, (cell_width_px - new_w) // 2)

                    ws.add_image(img_obj, f"{col_letter}{row_idx}")
                except Exception as e:
                    print(f"  [!] Embed poster failed for #{m['id']}: {e}")

    ws.auto_filter.ref = f"A1:H{len(movies) + 1}"
    ws.freeze_panes = "C2"

    wb.save(excel_path)
    print(f"[+] Saved Excel to {excel_path}")


def print_table(movies):
    if not movies:
        return

    display_cols = COLUMNS
    col_widths = {col: len(col) for col in display_cols}
    for m in movies:
        for col in display_cols:
            val = str(m.get(col, ""))
            display = val[:40] + "..." if len(val) > 40 else val
            col_widths[col] = max(col_widths[col], len(display))

    sep = "+" + "+".join("-" * (w + 2) for w in col_widths.values()) + "+"
    header = "|" + "|".join(f" {col.upper():<{col_widths[col]}} " for col in display_cols) + "|"

    print(f"\n{sep}")
    print(header)
    print(sep)

    for m in movies[:10]:
        row = "|"
        for col in display_cols:
            val = str(m.get(col, ""))
            display = val[:40] + "..." if len(val) > 40 else val
            row += f" {display:<{col_widths[col]}} |"
        print(row)

    if len(movies) > 10:
        print(f"  ... and {len(movies) - 10} more rows")
    print(sep)
    print(f"  Total: {len(movies)} movies")
    print(f"  Posters: {POSTER_DIR}")
    print(f"  SQLite:  {OUTPUT_DIR / 'movie.db'}")
    print(f"  Excel:   {OUTPUT_DIR / 'movies.xlsx'}\n")


if __name__ == "__main__":
    init_db()
    movies = crawl_all()
    if movies:
        movies = download_posters(movies)
        save_to_db(movies)
        save_csv(movies)
        save_excel(movies)
        print_table(movies)
